from __future__ import annotations

import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_FLOOR, ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from services.mabang import config as mabang_settings
from services.mabang.auth_constants import (
    PRIVATE_AMZ_HOST,
)
from services.mabang.export_common import configured_text as _configured_text
from shared.infra.net import erp_http_session, external_http_session
from shared.datasets import dataset_dir

from .combo_sku import ComboComponent, ComboSku, fetch_inventory_combos, normalize_sku_key

from ...auth import get_auth_context, refresh_mabang_auth
from ...cookies import build_cookie_header
from ...errors import MabangAuthError, MabangBusinessError, MabangRequestError

DEFAULT_STORE_MSKU_DIR = dataset_dir("replenish_store_msku")
DEFAULT_OUTPUT_DIR = dataset_dir("replenish_actual_inventory")
DEFAULT_WAREHOUSE_SEARCH_URL = "https://private-amz.mabangerp.com/index.php?mod=warehouse.searchwarehousestock"
DEFAULT_WAREHOUSE_EXPORT_URL = (
    "https://private-amz.mabangerp.com/index.php?mod=warehouse.doexportwarehousestock&flag=1&showRmbColumn=0"
)
DEFAULT_PRIVATE_AMZ_ORIGIN = "https://private-amz.mabangerp.com"
DEFAULT_PRIVATE_AMZ_REFERER = "https://private-amz.mabangerp.com/"
SOURCE = "mabang_store_msku_shenzhen_warehouse_inventory"
EXCEL_ROW_HEIGHT = 15
EXCEL_COLUMN_WIDTH = 15
WAREHOUSE_ID = "1014318"
WAREHOUSE_NAME = "深圳仓库"
STORE_MSKU_FILE_SUFFIXES = ("店铺MSKU数据", "msku_data")
ACTUAL_INVENTORY_FILE_SUFFIX = "真实库存（深圳仓库）"
ACTUAL_INVENTORY_QUANTITY_COLUMN = "真实库存（深圳仓库）数量"
COMBO_ACTUAL_INVENTORY_SHEET = "真实库存（深圳仓库）-组合sku"
STOCK_ACTUAL_INVENTORY_SHEET = "真实库存（深圳仓库）-库存sku"
NO_LOCAL_SKU_SHEET = "无本地SKU"
NO_INVENTORY_DATA_SHEET = "无库存数据"
SOURCE_FILE_RE = re.compile(
    rf"^(?P<source_time>\d{{12}})-(?P<store>.+)_(?:{'|'.join(re.escape(suffix) for suffix in STORE_MSKU_FILE_SUFFIXES)})\.xlsx$",
    re.IGNORECASE,
)
AUTH_FAIL_STATUS = {401, 403}
SALES_COLUMNS = ("7天销量", "14天销量", "30天销量")
FBA_STOCK_COLUMNS = ("可售", "待入库", "预留", "在途", "待调仓", "调仓中")
SOURCE_COLUMNS = ("MSKU", "父ASIN", "ASIN", "本地SKU", "商品链接", *SALES_COLUMNS, *FBA_STOCK_COLUMNS)
SOURCE_LOCAL_SKU_NAME_COLUMN = "本地SKU名称"
SOURCE_PRODUCT_NAME_COLUMN = "产品名称"
SOURCE_REMARK_COLUMN = "备注"
STOCK_SKU_COLUMN = "库存SKU编号"
AVAILABLE_STOCK_COLUMN = "可用库存量"
BASE_OUTPUT_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    SOURCE_LOCAL_SKU_NAME_COLUMN,
    SOURCE_PRODUCT_NAME_COLUMN,
    SOURCE_REMARK_COLUMN,
    "商品链接",
    ACTUAL_INVENTORY_QUANTITY_COLUMN,
    "子SKU",
)
INVENTORY_OUTPUT_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    SOURCE_LOCAL_SKU_NAME_COLUMN,
    SOURCE_PRODUCT_NAME_COLUMN,
    SOURCE_REMARK_COLUMN,
    "商品链接",
    "FBA总库存",
    "加权日销",
    "可销售天数",
    ACTUAL_INVENTORY_QUANTITY_COLUMN,
    "子SKU",
)
OUTPUT_COLUMNS = BASE_OUTPUT_COLUMNS
TWO_DECIMAL_COLUMNS = {"加权日销", "可销售天数"}
ACTUAL_INVENTORY_HIGHLIGHT_COLOR = "FFF2CC"
INVENTORY_HIGHLIGHT_COLUMNS = {"MSKU", ACTUAL_INVENTORY_QUANTITY_COLUMN}


class StoreMskuActualInventoryError(MabangBusinessError):
    pass


class StoreMskuActualInventoryAuthError(StoreMskuActualInventoryError, MabangAuthError):
    pass


@dataclass(frozen=True)
class SourceMskuFile:
    path: Path
    source_data_time: str
    source_datetime: datetime


@dataclass(frozen=True)
class StoreMskuRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    product_link: str
    local_sku_name: str = ""
    product_name: str = ""
    remark: str = ""
    sales_7d: Decimal = Decimal("0")
    sales_14d: Decimal = Decimal("0")
    sales_30d: Decimal = Decimal("0")
    fba_sellable: Decimal = Decimal("0")
    fba_inbound: Decimal = Decimal("0")
    fba_reserved: Decimal = Decimal("0")
    fba_in_transit: Decimal = Decimal("0")
    fba_pending_transfer: Decimal = Decimal("0")
    fba_transferring: Decimal = Decimal("0")


@dataclass(frozen=True)
class ActualInventoryRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    product_link: str
    actual_inventory: Decimal | None
    child_skus: str
    is_combo_sku: bool = False
    fba_total_inventory: Decimal = Decimal("0")
    weighted_daily_sales: Decimal = Decimal("0")
    sales_days: Decimal | None = None
    local_sku_name: str = ""
    product_name: str = ""
    remark: str = ""


@dataclass(frozen=True)
class ActualInventoryRowGroups:
    combo_inventory_rows: list[ActualInventoryRow]
    stock_inventory_rows: list[ActualInventoryRow]
    no_local_sku_rows: list[ActualInventoryRow]
    no_inventory_rows: list[ActualInventoryRow]

    @property
    def inventory_rows(self) -> list[ActualInventoryRow]:
        return [*self.combo_inventory_rows, *self.stock_inventory_rows]


@dataclass(frozen=True)
class ActualInventoryResult:
    store_name: str
    source_msku_xlsx_path: str
    source_msku_data_time: str
    unique_local_sku_count: int
    detected_combo_sku_count: int
    queried_warehouse_stock_sku_count: int
    missing_warehouse_stock_skus: list[str]
    shenzhen_warehouse_inventory_report_xlsx_path: str
    matched_warehouse_inventory_msku_row_count: int = 0
    missing_local_sku_msku_row_count: int = 0
    missing_warehouse_inventory_msku_row_count: int = 0
    warehouse_id: str = WAREHOUSE_ID
    warehouse_name: str = WAREHOUSE_NAME
    result_source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "warehouse_id": self.warehouse_id,
            "warehouse_name": self.warehouse_name,
            "source_msku_xlsx_path": self.source_msku_xlsx_path,
            "source_msku_data_time": self.source_msku_data_time,
            "unique_local_sku_count": self.unique_local_sku_count,
            "detected_combo_sku_count": self.detected_combo_sku_count,
            "queried_warehouse_stock_sku_count": self.queried_warehouse_stock_sku_count,
            "matched_warehouse_inventory_msku_row_count": self.matched_warehouse_inventory_msku_row_count,
            "missing_local_sku_msku_row_count": self.missing_local_sku_msku_row_count,
            "missing_warehouse_inventory_msku_row_count": self.missing_warehouse_inventory_msku_row_count,
            "missing_warehouse_stock_sku_count": len(self.missing_warehouse_stock_skus),
            "missing_warehouse_stock_skus": list(self.missing_warehouse_stock_skus),
            "shenzhen_warehouse_inventory_report_xlsx_path": self.shenzhen_warehouse_inventory_report_xlsx_path,
            "result_source": self.result_source,
        }


def _clean_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "nan":
        return ""
    return text


def _safe_file_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("._-") or "store_msku_inventory"


def _timestamp_text(value: datetime | None = None) -> str:
    return (value or datetime.now()).strftime("%Y%m%d%H%M")


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _unique_text(values: list[str] | tuple[str, ...] | Any) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for value in values or []:
        text = _clean_text(value)
        key = normalize_sku_key(text)
        if not key or key in unique:
            continue
        unique[key] = text
    return list(unique.values())


def _decimal_value(value: Any, *, default: Decimal = Decimal("0")) -> Decimal:
    text = _clean_text(value).replace(",", "")
    if not text:
        return default
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def _display_decimal(value: Decimal | None) -> float | int | str:
    if value is None:
        return ""
    normalized = value.normalize()
    if normalized == normalized.to_integral_value():
        return int(normalized)
    return float(normalized)


def _display_two_decimal(value: Decimal | None) -> float | str:
    if value is None:
        return ""
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _display_quantity(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral_value():
        return str(int(normalized))
    return format(normalized, "f")


def _resolve_store_msku_dir(input_dir: str | Path | None = None) -> Path:
    if input_dir is not None:
        return Path(input_dir)
    configured = str(mabang_settings.MABANG_STORE_MSKU_OUTPUT_DIR or "").strip()
    return Path(configured) if configured else DEFAULT_STORE_MSKU_DIR


def _resolve_output_dir(output_dir: str | Path | None = None) -> Path:
    if output_dir is not None:
        path = Path(output_dir)
    else:
        configured = str(mabang_settings.MABANG_STORE_MSKU_INVENTORY_OUTPUT_DIR or "").strip()
        path = Path(configured) if configured else DEFAULT_OUTPUT_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def find_latest_store_msku_file(store_name: str, *, input_dir: str | Path | None = None) -> SourceMskuFile:
    clean_store_name = normalize_store_name(store_name)
    directory = _resolve_store_msku_dir(input_dir)
    safe_store_name = _safe_file_part(clean_store_name)
    if not directory.is_dir():
        raise StoreMskuActualInventoryError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")

    candidates: list[SourceMskuFile] = []
    for path in directory.glob(f"*-{safe_store_name}_*.xlsx"):
        match = SOURCE_FILE_RE.match(path.name)
        if not match:
            continue
        source_data_time = match.group("source_time")
        try:
            source_datetime = datetime.strptime(source_data_time, "%Y%m%d%H%M")
        except ValueError:
            continue
        candidates.append(SourceMskuFile(path=path, source_data_time=source_data_time, source_datetime=source_datetime))
    if not candidates:
        raise StoreMskuActualInventoryError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")
    return max(candidates, key=lambda item: (item.source_datetime, item.path.name))


def load_store_msku_rows(xlsx_path: str | Path) -> list[StoreMskuRow]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        header_values = next(values, None)
        headers = [_clean_text(cell) for cell in list(header_values or [])]
        missing = [column for column in SOURCE_COLUMNS if column not in headers]
        if missing:
            raise StoreMskuActualInventoryError(f"店铺MSKU数据缺少列: {', '.join(missing)}")

        rows: list[StoreMskuRow] = []
        for row_values in values:
            row = dict(zip(headers, list(row_values or []), strict=False))
            if not any(_clean_text(value) for value in row.values()):
                continue
            rows.append(
                StoreMskuRow(
                    msku=_clean_text(row.get("MSKU")),
                    parent_asin=_clean_text(row.get("父ASIN")),
                    asin=_clean_text(row.get("ASIN")),
                    local_sku=_clean_text(row.get("本地SKU")),
                    product_link=_clean_text(row.get("商品链接")),
                    local_sku_name=_clean_text(row.get(SOURCE_LOCAL_SKU_NAME_COLUMN)),
                    product_name=_clean_text(row.get(SOURCE_PRODUCT_NAME_COLUMN)),
                    remark=_clean_text(row.get(SOURCE_REMARK_COLUMN)),
                    sales_7d=_decimal_value(row.get("7天销量")),
                    sales_14d=_decimal_value(row.get("14天销量")),
                    sales_30d=_decimal_value(row.get("30天销量")),
                    fba_sellable=_decimal_value(row.get("可售")),
                    fba_inbound=_decimal_value(row.get("待入库")),
                    fba_reserved=_decimal_value(row.get("预留")),
                    fba_in_transit=_decimal_value(row.get("在途")),
                    fba_pending_transfer=_decimal_value(row.get("待调仓")),
                    fba_transferring=_decimal_value(row.get("调仓中")),
                )
            )
    except StoreMskuActualInventoryError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    return rows


def _private_amz_post_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": _configured_text("MABANG_WAREHOUSE_STOCK_ORIGIN", DEFAULT_PRIVATE_AMZ_ORIGIN),
        "Referer": _configured_text("MABANG_WAREHOUSE_STOCK_REFERER", DEFAULT_PRIVATE_AMZ_REFERER),
        "Cookie": cookie_header,
    }


def _private_amz_get_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "Cookie": cookie_header,
    }


async def _read_json_response(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuActualInventoryAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")

    data: Any = None
    if text:
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            data = None
    if data is None:
        try:
            data = await resp.json(content_type=None)
        except Exception:
            data = None
    if not isinstance(data, dict):
        raise StoreMskuActualInventoryError(f"{action}返回非JSON对象")
    if data.get("success") is False:
        message = _clean_text(data.get("msg") or data.get("message") or data.get("error") or "unknown")
        raise StoreMskuActualInventoryError(f"{action}业务异常: {message}")
    return data


async def _read_optional_json_response(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuActualInventoryAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")
    if not text:
        return {}
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return {}
    if isinstance(data, dict) and data.get("success") is False:
        message = _clean_text(data.get("msg") or data.get("message") or data.get("error") or "unknown")
        raise StoreMskuActualInventoryError(f"{action}业务异常: {message}")
    return data if isinstance(data, dict) else {}


async def _resolve_private_amz_cookie() -> str:
    context = await get_auth_context(purpose="store_msku_actual_inventory_download")
    cookie_header = build_cookie_header(
        context.cookies_by_domain,
        request_host=PRIVATE_AMZ_HOST,
        extra_cookies={"mabang_lite_rowsPerPage": "100"},
    )
    if not cookie_header:
        raise StoreMskuActualInventoryAuthError("未获取到 private-amz.mabangerp.com Cookie")
    return cookie_header


def _warehouse_search_url() -> str:
    return _configured_text("MABANG_WAREHOUSE_STOCK_SEARCH_URL", DEFAULT_WAREHOUSE_SEARCH_URL)


def _warehouse_export_url() -> str:
    return _configured_text("MABANG_WAREHOUSE_STOCK_EXPORT_URL", DEFAULT_WAREHOUSE_EXPORT_URL)


def _ids_lines(values: list[str]) -> str:
    return "\r\n".join(values) + "\r\n"


async def _download_xlsx_from_url(file_url: str, target_path: Path, *, action: str) -> Path:
    url = _clean_text(file_url)
    if not url:
        raise ValueError("file_url 不能为空")
    target_path.parent.mkdir(parents=True, exist_ok=True)
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*"
    }
    async with external_http_session.get(url, headers=headers) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"{action}失败(status={status_code}): {msg}")
        if not body:
            raise StoreMskuActualInventoryError(f"{action}返回空文件")
    target_path.write_bytes(body)
    return target_path


def _warehouse_search_form_data(stock_skus: list[str]) -> list[tuple[str, str]]:
    return [
        ("stockOrderby", ""),
        ("parentCategoryId", ""),
        ("categoryId", ""),
        ("third_category_id", ""),
        ("warehouseIds[]", WAREHOUSE_ID),
        ("stockName", "nameCN"),
        ("stockNameValue", ""),
        ("statusIN[]", "3"),
        ("inventoryAlertId", "0"),
        ("livenessType", ""),
        ("isNewType", ""),
        ("gridcodeStr", ""),
        ("stockSkuStr", _ids_lines(stock_skus)),
        ("page", "1"),
        ("rowsPerPage", "50"),
        ("warehouseId", "undefined"),
        ("startTime", ""),
        ("endTime", ""),
        ("isIdn", "1"),
        ("warehouseIdArr", ""),
        ("stockQuantitylt", ""),
        ("stockQuantitygt", ""),
        ("stockWarningQuantitylt", ""),
        ("stockWarningQuantitygt", ""),
        ("saleAvailableDayslt", ""),
        ("saleAvailableDaysgt", ""),
    ]


async def search_warehouse_stock(stock_skus: list[str]) -> None:
    if not stock_skus:
        return
    cookie_header = await _resolve_private_amz_cookie()
    async with erp_http_session.post(
        _warehouse_search_url(),
        data=_warehouse_search_form_data(stock_skus),
        headers=_private_amz_post_headers(cookie_header),
    ) as resp:
        await _read_optional_json_response(resp, action="库存SKU搜索")


async def download_warehouse_stock_xlsx(
    *,
    store_name: str,
    output_dir: str | Path | None = None,
) -> Path:
    cookie_header = await _resolve_private_amz_cookie()
    directory = _resolve_output_dir(output_dir)
    target_path = directory / f"{_timestamp_text()}-{_safe_file_part(store_name)}_warehouse_stock.xlsx"
    async with erp_http_session.get(
        _warehouse_export_url(),
        headers=_private_amz_get_headers(cookie_header),
    ) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code in AUTH_FAIL_STATUS:
            raise StoreMskuActualInventoryAuthError(f"库存SKU导出鉴权失败(status={status_code})")
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"库存SKU导出请求失败(status={status_code}): {msg}")
        if not body:
            raise StoreMskuActualInventoryError("库存SKU导出返回空文件")
    target_path.write_bytes(body)
    return target_path


def parse_stock_inventory_xlsx(xlsx_path: str | Path) -> dict[str, Decimal]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取库存SKU导出xlsx") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"库存SKU导出xlsx不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        missing = [column for column in (STOCK_SKU_COLUMN, AVAILABLE_STOCK_COLUMN) if column not in headers]
        if missing:
            raise StoreMskuActualInventoryError(f"库存SKU导出xlsx缺少列: {', '.join(missing)}")

        quantities: dict[str, Decimal] = {}
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            stock_sku = _clean_text(row.get(STOCK_SKU_COLUMN))
            key = normalize_sku_key(stock_sku)
            if not key:
                continue
            quantities[key] = quantities.get(key, Decimal("0")) + _decimal_value(row.get(AVAILABLE_STOCK_COLUMN))
    except StoreMskuActualInventoryError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取库存SKU导出xlsx失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass
    return quantities


def stock_skus_for_inventory(local_skus: list[str], combo_map: dict[str, ComboSku]) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for local_sku in local_skus:
        key = normalize_sku_key(local_sku)
        if not key:
            continue
        combo = combo_map.get(key)
        if combo is None:
            unique.setdefault(key, local_sku)
            continue
        for component in combo.components:
            component_key = normalize_sku_key(component.stock_sku)
            if component_key:
                unique.setdefault(component_key, component.stock_sku)
    return list(unique.values())


def _combo_child_text(combo: ComboSku) -> str:
    return ", ".join(f"{component.stock_sku} * {_display_quantity(component.quantity)}" for component in combo.components)


def _fba_total_inventory(row: StoreMskuRow) -> Decimal:
    return (
        row.fba_sellable
        + row.fba_inbound
        + row.fba_reserved
        + row.fba_in_transit
        + row.fba_pending_transfer
        + row.fba_transferring
    )


def _weighted_daily_sales(row: StoreMskuRow) -> Decimal:
    return (
        row.sales_7d / Decimal("7") * Decimal("0.6")
        + row.sales_14d / Decimal("14") * Decimal("0.3")
        + row.sales_30d / Decimal("30") * Decimal("0.1")
    )


def _sales_days(fba_total_inventory: Decimal, weighted_daily_sales: Decimal) -> Decimal | None:
    if weighted_daily_sales == 0:
        return None
    return fba_total_inventory / weighted_daily_sales


def calculate_inventory_rows(
    msku_rows: list[StoreMskuRow],
    *,
    combo_map: dict[str, ComboSku],
    stock_quantities: dict[str, Decimal],
) -> tuple[list[ActualInventoryRow], list[str]]:
    missing: OrderedDict[str, str] = OrderedDict()
    result_rows: list[ActualInventoryRow] = []
    for row in msku_rows:
        local_key = normalize_sku_key(row.local_sku)
        combo = combo_map.get(local_key)
        is_combo_sku = combo is not None
        fba_total_inventory = _fba_total_inventory(row)
        weighted_daily_sales = _weighted_daily_sales(row)
        actual_inventory: Decimal | None
        child_skus = ""
        if not local_key:
            actual_inventory = None
        elif combo is None:
            actual_inventory = stock_quantities.get(local_key)
            if actual_inventory is None:
                missing.setdefault(local_key, row.local_sku)
        else:
            child_skus = _combo_child_text(combo)
            possible_counts: list[Decimal] = []
            for component in combo.components:
                component_key = normalize_sku_key(component.stock_sku)
                quantity = stock_quantities.get(component_key)
                if quantity is None:
                    missing.setdefault(component_key, component.stock_sku)
                    continue
                possible_counts.append((quantity / component.quantity).to_integral_value(rounding=ROUND_FLOOR))
            actual_inventory = min(possible_counts) if len(possible_counts) == len(combo.components) else None

        result_rows.append(
            ActualInventoryRow(
                msku=row.msku,
                parent_asin=row.parent_asin,
                asin=row.asin,
                local_sku=row.local_sku,
                product_link=row.product_link,
                actual_inventory=actual_inventory,
                child_skus=child_skus,
                is_combo_sku=is_combo_sku,
                fba_total_inventory=fba_total_inventory,
                weighted_daily_sales=weighted_daily_sales,
                sales_days=_sales_days(fba_total_inventory, weighted_daily_sales),
                local_sku_name=row.local_sku_name,
                product_name=row.product_name,
                remark=row.remark,
            )
        )
    return result_rows, list(missing.values())


def split_inventory_rows(rows: list[ActualInventoryRow]) -> ActualInventoryRowGroups:
    combo_inventory_rows: list[ActualInventoryRow] = []
    stock_inventory_rows: list[ActualInventoryRow] = []
    no_local_sku_rows: list[ActualInventoryRow] = []
    no_inventory_rows: list[ActualInventoryRow] = []
    for row in rows:
        if not normalize_sku_key(row.local_sku):
            no_local_sku_rows.append(row)
        elif row.actual_inventory is None:
            no_inventory_rows.append(row)
        elif row.is_combo_sku:
            combo_inventory_rows.append(row)
        else:
            stock_inventory_rows.append(row)
    return ActualInventoryRowGroups(
        combo_inventory_rows=combo_inventory_rows,
        stock_inventory_rows=stock_inventory_rows,
        no_local_sku_rows=no_local_sku_rows,
        no_inventory_rows=no_inventory_rows,
    )


def _base_row_values(row: ActualInventoryRow) -> list[Any]:
    return [
        row.msku,
        row.parent_asin,
        row.asin,
        row.local_sku,
        row.local_sku_name,
        row.product_name,
        row.remark,
        row.product_link,
        _display_decimal(row.actual_inventory),
        row.child_skus,
    ]


def _inventory_row_values(row: ActualInventoryRow) -> list[Any]:
    return [
        row.msku,
        row.parent_asin,
        row.asin,
        row.local_sku,
        row.local_sku_name,
        row.product_name,
        row.remark,
        row.product_link,
        _display_decimal(row.fba_total_inventory),
        _display_two_decimal(row.weighted_daily_sales),
        _display_two_decimal(row.sales_days),
        _display_decimal(row.actual_inventory),
        row.child_skus,
    ]


def _sorted_inventory_rows(rows: list[ActualInventoryRow]) -> list[ActualInventoryRow]:
    return sorted(rows, key=lambda row: row.weighted_daily_sales, reverse=True)


def _append_inventory_sheet(
    workbook: Any,
    title: str,
    columns: tuple[str, ...],
    rows: list[ActualInventoryRow],
    row_values: Any,
    *,
    active: bool = False,
    highlight_actual_inventory: bool = False,
) -> None:
    from openpyxl.styles import PatternFill

    worksheet = workbook.active if active else workbook.create_sheet(title)
    worksheet.title = title
    worksheet.append(list(columns))
    for row in rows:
        worksheet.append(row_values(row))
    for index, header in enumerate(columns, start=1):
        if header not in TWO_DECIMAL_COLUMNS:
            continue
        for cells in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cells[0].number_format = "0.00"
    if highlight_actual_inventory:
        fill = PatternFill(fill_type="solid", fgColor=ACTUAL_INVENTORY_HIGHLIGHT_COLOR)
        for column_index, header in enumerate(columns, start=1):
            if header not in INVENTORY_HIGHLIGHT_COLUMNS:
                continue
            for cells in worksheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
                cells[0].fill = fill
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = EXCEL_ROW_HEIGHT
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = EXCEL_COLUMN_WIDTH


def write_actual_inventory_xlsx(rows: list[ActualInventoryRow], output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入真实库存（深圳仓库）xlsx") from exc

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        groups = split_inventory_rows(rows)
        _append_inventory_sheet(
            workbook,
            COMBO_ACTUAL_INVENTORY_SHEET,
            INVENTORY_OUTPUT_COLUMNS,
            _sorted_inventory_rows(groups.combo_inventory_rows),
            _inventory_row_values,
            active=True,
            highlight_actual_inventory=True,
        )
        _append_inventory_sheet(
            workbook,
            STOCK_ACTUAL_INVENTORY_SHEET,
            INVENTORY_OUTPUT_COLUMNS,
            _sorted_inventory_rows(groups.stock_inventory_rows),
            _inventory_row_values,
            highlight_actual_inventory=True,
        )
        _append_inventory_sheet(workbook, NO_LOCAL_SKU_SHEET, BASE_OUTPUT_COLUMNS, groups.no_local_sku_rows, _base_row_values)
        _append_inventory_sheet(workbook, NO_INVENTORY_DATA_SHEET, BASE_OUTPUT_COLUMNS, groups.no_inventory_rows, _base_row_values)
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


async def _export_store_msku_actual_inventory_once(
    store_name: str,
    *,
    input_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> ActualInventoryResult:
    clean_store_name = normalize_store_name(store_name)
    source = find_latest_store_msku_file(clean_store_name, input_dir=input_dir)
    msku_rows = load_store_msku_rows(source.path)
    local_skus = _unique_text([row.local_sku for row in msku_rows])

    output_directory = _resolve_output_dir(output_dir)
    combo_map = await fetch_inventory_combos(clean_store_name, msku_rows)
    stock_skus = stock_skus_for_inventory(local_skus, combo_map)

    async def warehouse_once() -> dict[str, Decimal]:
        await search_warehouse_stock(stock_skus)
        stock_xlsx_path = await download_warehouse_stock_xlsx(
            store_name=clean_store_name, output_dir=output_directory,
        )
        return parse_stock_inventory_xlsx(stock_xlsx_path)

    try:
        stock_quantities = await warehouse_once()
    except StoreMskuActualInventoryAuthError:
        await refresh_mabang_auth(purpose="store_msku_actual_inventory_auth_retry")
        stock_quantities = await warehouse_once()
    inventory_rows, missing_stock_skus = calculate_inventory_rows(
        msku_rows,
        combo_map=combo_map,
        stock_quantities=stock_quantities,
    )
    inventory_groups = split_inventory_rows(inventory_rows)

    final_xlsx_path = output_directory / f"{source.source_data_time}-{_safe_file_part(clean_store_name)}_{ACTUAL_INVENTORY_FILE_SUFFIX}.xlsx"
    write_actual_inventory_xlsx(inventory_rows, final_xlsx_path)
    return ActualInventoryResult(
        store_name=clean_store_name,
        source_msku_xlsx_path=str(source.path),
        source_msku_data_time=source.source_data_time,
        unique_local_sku_count=len(local_skus),
        detected_combo_sku_count=len(combo_map),
        queried_warehouse_stock_sku_count=len(stock_skus),
        missing_warehouse_stock_skus=missing_stock_skus,
        shenzhen_warehouse_inventory_report_xlsx_path=str(final_xlsx_path),
        matched_warehouse_inventory_msku_row_count=len(inventory_groups.inventory_rows),
        missing_local_sku_msku_row_count=len(inventory_groups.no_local_sku_rows),
        missing_warehouse_inventory_msku_row_count=len(inventory_groups.no_inventory_rows),
    )


async def export_store_msku_actual_inventory(
    store_name: str,
    *,
    input_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> ActualInventoryResult:
    return await _export_store_msku_actual_inventory_once(
        store_name, input_dir=input_dir, output_dir=output_dir,
    )


__all__ = [
    "ComboSku",
    "ComboComponent",
    "normalize_sku_key",
    "AVAILABLE_STOCK_COLUMN",
    "ACTUAL_INVENTORY_FILE_SUFFIX",
    "BASE_OUTPUT_COLUMNS",
    "FBA_STOCK_COLUMNS",
    "INVENTORY_OUTPUT_COLUMNS",
    "OUTPUT_COLUMNS",
    "SALES_COLUMNS",
    "SOURCE",
    "STOCK_SKU_COLUMN",
    "STORE_MSKU_FILE_SUFFIXES",
    "WAREHOUSE_ID",
    "WAREHOUSE_NAME",
    "ActualInventoryResult",
    "ActualInventoryRow",
    "ActualInventoryRowGroups",
    "StoreMskuActualInventoryAuthError",
    "StoreMskuActualInventoryError",
    "StoreMskuRow",
    "calculate_inventory_rows",
    "export_store_msku_actual_inventory",
    "find_latest_store_msku_file",
    "load_store_msku_rows",
    "parse_stock_inventory_xlsx",
    "search_warehouse_stock",
    "split_inventory_rows",
    "stock_skus_for_inventory",
    "write_actual_inventory_xlsx",
    "_timestamp_text",
    "_warehouse_search_form_data",
]
